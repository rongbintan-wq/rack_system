"""Excel 设备导入：openpyxl 解析 + U 位冲突检测 + 增量更新 + 事务回滚。

列映射（固定顺序，表头在首行）：
A 资源ID | B 资源编号 | C 设备类型 | D 品牌名称 | E 型号 | F 区域省份城市场地
G 机房名称 | H 机柜名称 | I 机柜编号 | J 起始U位 | K 占用U位数 | L 资产状态
M SN序列号 | N 主机名

冲突检测核心（修正 off-by-one 与排除自身 UPDATE）：
    区间交集：max(start_a, start_b) <= min(end_a, end_b)
    - 仅比对 is_deleted=False 的设备
    - 同 batch 内「资源编号」将原地 UPDATE 的旧记录须从比对集合排除
    - 同时检测 batch 内多行之间的相互重叠
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import crud
from app.models import Device, Rack

REQUIRED_COLS = {
    "A": "resource_id",
    "B": "resource_code",
    "C": "device_type",
    "D": "brand_name",
    "E": "model",
    "F": "region",
    "G": "room_name",
    "H": "rack_name",
    "I": "rack_code",
    "J": "start_u",
    "K": "height_u",
    "L": "asset_status",
}


@dataclass
class RawRow:
    row: int
    data: dict
    errors: list = field(default_factory=list)


def _cell(values, idx):
    if idx >= len(values):
        return None
    v = values[idx]
    return v.strip() if isinstance(v, str) else v


def extract_rows(file_bytes: bytes) -> list[RawRow]:
    """读取 .xlsx，跳过表头，逐行解析为 RawRow（仅做必填与类型校验）。"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows: list[RawRow] = []
    for i, values in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue  # 表头
        if values is None or all(v is None for v in values):
            continue
        rr = RawRow(row=i, data={})
        # A-N 共 14 列
        rr.data["resource_id"] = _cell(values, 0) or ""
        rr.data["resource_code"] = _cell(values, 1)
        rr.data["device_type"] = _cell(values, 2)
        rr.data["brand_name"] = _cell(values, 3) or ""
        rr.data["model"] = _cell(values, 4) or ""
        rr.data["region"] = _cell(values, 5) or ""
        rr.data["room_name"] = _cell(values, 6) or ""
        rr.data["rack_name"] = _cell(values, 7) or ""
        rr.data["rack_code"] = _cell(values, 8)
        rr.data["start_u"] = _cell(values, 9)
        rr.data["height_u"] = _cell(values, 10)
        rr.data["asset_status"] = _cell(values, 11) or "运行中"
        rr.data["sn"] = _cell(values, 12) or ""
        rr.data["hostname"] = _cell(values, 13) or ""

        # 必填校验
        for key in ["resource_code", "device_type", "rack_code", "start_u", "height_u"]:
            if rr.data.get(key) in (None, ""):
                rr.errors.append(f"第 {i} 行：必填列缺失（{key}）")
        # 整数校验
        for key in ["start_u", "height_u"]:
            v = rr.data.get(key)
            if v not in (None, ""):
                try:
                    rr.data[key] = int(v)
                except (ValueError, TypeError):
                    rr.errors.append(f"第 {i} 行：{key} 必须为整数，实际值「{v}」")
        if rr.data.get("height_u") is not None and isinstance(rr.data.get("height_u"), int):
            if rr.data["height_u"] < 1:
                rr.errors.append(f"第 {i} 行：占用U位数必须 >= 1")
        rows.append(rr)
    return rows


@dataclass
class ValidRow:
    row: int
    data: dict
    action: str  # insert | update
    existing_id: Optional[int] = None
    rack: Optional[Rack] = None


def validate_rows(db: Session, raw_rows: list[RawRow]):
    """返回 (valid_rows, errors)。valid_rows 已确定 insert/update 及目标机柜。"""
    valid: list[ValidRow] = []
    errors: list[dict] = []

    # 先收集本批 resource_code 集合，用于排除自身 UPDATE
    batch_codes = {r.data.get("resource_code") for r in raw_rows if r.data.get("resource_code")}

    # 按机柜分组的逐行冲突检测工作集
    # 结构：rack_code -> {"db_free": [intervals 来自DB且非本批更新], "tentative": [已接受batch区间]}
    rack_workset: dict[str, dict] = {}

    for rr in raw_rows:
        if rr.errors:
            for e in rr.errors:
                errors.append({"row": rr.row, "resource_code": rr.data.get("resource_code", ""), "reason": e})
            continue

        rc = rr.data["rack_code"]
        # 机柜归属校验
        rack = crud.get_rack_by_code(db, rc)
        if not rack:
            errors.append({"row": rr.row, "resource_code": rc, "reason": f"第 {rr.row} 行：机柜编号 {rc} 不存在，请先新增该机柜"})
            continue
        if rr.data["room_name"] and rack.room and rack.room.room_name != rr.data["room_name"]:
            errors.append({"row": rr.row, "resource_code": rc, "reason": f"第 {rr.row} 行：机柜编号 {rc} 不属于机房「{rr.data['room_name']}」"})
            continue

        start_u = rr.data["start_u"]
        height_u = rr.data["height_u"]
        end_u = start_u + height_u - 1

        # U 越界校验
        if start_u < 1 or end_u > rack.height_u:
            errors.append({"row": rr.row, "resource_code": rc, "reason": f"第 {rr.row} 行：起始U位 {start_u} + 占用U数 {height_u} 超出机柜总高度 {rack.height_u}U"})
            continue

        # 初始化工作集
        if rc not in rack_workset:
            # DB 中该机柜的已占用区间，排除本批将 UPDATE 的旧记录
            db_intervals = []
            for d in db.query(Device).filter(Device.rack_code == rc, Device.is_deleted.is_(False)):
                if d.resource_code in batch_codes:
                    continue  # 自身旧记录排除
                db_intervals.append((d.start_u, d.start_u + d.height_u - 1, d.resource_code))
            rack_workset[rc] = {"db": db_intervals, "tentative": []}

        ws = rack_workset[rc]
        # 冲突检测：与新区间 [start_u, end_u]
        conflict_with = None
        for (s, e, code) in ws["db"]:
            if max(start_u, s) <= min(end_u, e):
                conflict_with = (s, e, code)
                break
        if not conflict_with:
            for (s, e, code) in ws["tentative"]:
                if max(start_u, s) <= min(end_u, e):
                    conflict_with = (s, e, code)
                    break

        if conflict_with:
            s, e, code = conflict_with
            errors.append({"row": rr.row, "resource_code": rc, "reason": f"第 {rr.row} 行：U 位 {start_u}-{end_u} 与资源编号 {code}（U {s}-{e}）冲突"})
            continue

        # 判定 insert / update
        existing = crud.get_device_by_resource_code(db, rr.data["resource_code"])
        action = "update" if existing else "insert"
        vr = ValidRow(row=rr.row, data=rr.data, action=action, existing_id=existing.id if existing else None, rack=rack)
        valid.append(vr)
        # 将本行区间加入 tentative，供后续同 batch 行检测
        ws["tentative"].append((start_u, end_u, rr.data["resource_code"]))

    return valid, errors


def apply_rows(db: Session, valid_rows: list[ValidRow], operator=None):
    """在事务内写入；调用方负责 commit/rollback。"""
    for vr in valid_rows:
        d = vr.data
        common = {
            "resource_id": d.get("resource_id", ""),
            "resource_code": d["resource_code"],
            "device_type": d["device_type"],
            "brand_name": d.get("brand_name", ""),
            "model": d.get("model", ""),
            "region": d.get("region", ""),
            "site_detail": d.get("room_name", ""),  # 场地详情冗余
            "room_name": d.get("room_name", ""),
            "rack_name": d.get("rack_name", ""),
            "rack_code": d["rack_code"],
            "start_u": d["start_u"],
            "height_u": d["height_u"],
            "asset_status": d.get("asset_status", "运行中"),
            "sn": d.get("sn", ""),
            "hostname": d.get("hostname", ""),
        }
        if vr.action == "update" and vr.existing_id:
            dev = crud.get_device(db, vr.existing_id)
            crud.update_device(db, dev, common, operator)
        else:
            crud.create_device(db, common, operator)
